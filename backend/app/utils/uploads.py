import io
import json
import logging
import zipfile

from fastapi import UploadFile
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from pydantic import ValidationError

from backend.common.core.enums import SiteStatus
from backend.common.models.site import BaseUrl, ScrapeMethodConfiguration, Site


def parse_line(line):
    name: str
    base_url_str: str
    tag_str: str
    doc_ext_str: str
    url_keyw_str: str
    collection_method: str
    scrape_method = "SimpleDocumentScrape"
    cron = "0 16 * * *"
    name, base_url_str, tag_str, doc_ext_str, url_keyw_str, collection_method = line  # type: ignore

    tags = tag_str.split(",") if tag_str else []
    base_urls = base_url_str.split("|") if base_url_str else []
    doc_exts = doc_ext_str.split(",") if doc_ext_str else ["pdf"]
    url_keyws = url_keyw_str.split(",") if url_keyw_str else []
    collection_method = collection_method if collection_method else "AUTOMATED"
    scrape_method_configuration = ScrapeMethodConfiguration(
        document_extensions=doc_exts,
        url_keywords=url_keyws,
        proxy_exclusions=[],
        wait_for=[],
        follow_links=False,
        follow_link_keywords=[],
        follow_link_url_keywords=[],
    )

    clean_urls = []
    for base_url in base_urls:
        try:
            parsed_url = BaseUrl(url=base_url)
            clean_urls.append(parsed_url)
        except ValidationError:
            logging.error(f"site {name} has invalid url: {base_url}")

    return Site(
        name=name,
        base_urls=clean_urls,
        scrape_method=scrape_method,
        collection_method=collection_method,
        scrape_method_configuration=scrape_method_configuration,
        tags=tags,
        disabled=False,
        cron=cron,
    )


def get_sites_from_json(file: UploadFile):
    content = file.file.read()
    content_obj = json.loads(content)
    for site in content_obj:
        site["disabled"] = False
        site["status"] = SiteStatus.NEW
        new_site = Site.parse_obj(site)
        yield new_site


def get_lines_from_xlsx(file: UploadFile):
    wb = load_workbook(io.BytesIO(file.file.read()))
    sheet = wb[wb.sheetnames[0]]

    for i, line in enumerate(sheet.values):
        # Skip header.
        if i == 0:
            continue
        # Skip blank site names. Happens with last line.
        if not line[0]:
            continue
        # Remove illegal characters.
        clean_line = []
        for line_value in line:
            if isinstance(line_value, str):
                clean_line.append(ILLEGAL_CHARACTERS_RE.sub("", line_value))
            else:
                clean_line.append(line_value)
        # Yield parsed site.
        yield parse_line(clean_line)


def get_lines_from_text_file(file: UploadFile):
    for line in file.file:
        line = line.decode("utf-8").strip()
        yield parse_line(line.split("\t"))


def get_sites_from_upload(file: UploadFile):
    if file.content_type == "application/json":
        return get_sites_from_json(file)
    try:
        return get_lines_from_xlsx(file)
    except zipfile.BadZipFile:
        return get_lines_from_text_file(file)
