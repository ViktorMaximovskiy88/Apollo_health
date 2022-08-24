import io
import json
import logging
import urllib.parse
import zipfile

from beanie import PydanticObjectId
from beanie.operators import ElemMatch
from fastapi import APIRouter, Depends, HTTPException, Query, Security, UploadFile, status
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from pydantic import BaseModel, ValidationError

from backend.app.routes.table_query import (
    TableFilterInfo,
    TableQueryResponse,
    TableSortInfo,
    get_query_json_list,
    query_table,
)
from backend.app.utils.logger import (
    Logger,
    create_and_log,
    get_logger,
    update_and_log_diff,
    update_and_log_diff_forground,
)
from backend.app.utils.user import get_current_user
from backend.common.core.enums import SiteStatus
from backend.common.models.site import (
    BaseUrl,
    NewSite,
    ScrapeMethodConfiguration,
    Site,
    UpdateSite,
    UpdateSiteAssigne,
)
from backend.common.models.site_scrape_task import SiteScrapeTask
from backend.common.models.user import User

router = APIRouter(
    prefix="/sites",
    tags=["Sites"],
)


async def get_target(id: PydanticObjectId) -> Site:
    site = await Site.get(id)
    if not site:
        raise HTTPException(detail=f"Site {id} Not Found", status_code=status.HTTP_404_NOT_FOUND)
    return site


@router.get("/", response_model=TableQueryResponse)
async def read_sites(
    current_user: User = Depends(get_current_user),
    limit: int | None = None,
    skip: int | None = None,
    sorts: list[TableSortInfo] = Depends(get_query_json_list("sorts", TableSortInfo)),
    filters: list[TableFilterInfo] = Depends(get_query_json_list("filters", TableFilterInfo)),
):
    query = Site.find({"disabled": False})
    return await query_table(query, limit, skip, sorts, filters)


@router.get("/download", response_model=list[NewSite], dependencies=[Security(get_current_user)])
async def download_sites():
    return (
        await Site.find({"disabled": False, "status": {"$ne": SiteStatus.INACTIVE}})
        .project(NewSite)
        .to_list()
    )


class ActiveUrlResponse(BaseModel):
    in_use: bool
    site: Site | None = None


@router.get(
    "/active-url",
    response_model=ActiveUrlResponse,
    dependencies=[Security(get_current_user)],
)
async def check_url(
    url: str,
    current_site: PydanticObjectId | None = Query(default=None, alias="currentSite"),
):
    site = await Site.find_one(
        ElemMatch(Site.base_urls, {"url": urllib.parse.unquote(url)}),
        Site.id != current_site,
        {"disabled": False},
    )

    if site:
        return ActiveUrlResponse(in_use=True, site=site)
    else:
        return ActiveUrlResponse(in_use=False)


@router.get("/{id}", response_model=Site, dependencies=[Security(get_current_user)])
async def read_site(
    target: Site = Depends(get_target),
):
    return target


@router.put("/", response_model=Site, status_code=status.HTTP_201_CREATED)
async def create_site(
    site: NewSite,
    current_user: User = Security(get_current_user),
    logger: Logger = Depends(get_logger),
):
    new_site = Site(
        name=site.name,
        creator_id=current_user.id,
        base_urls=site.base_urls,
        scrape_method=site.scrape_method,
        collection_method=site.collection_method,
        scrape_method_configuration=site.scrape_method_configuration,
        tags=site.tags,
        disabled=False,
        cron=site.cron,
    )
    await create_and_log(logger, current_user, new_site)
    return new_site


def parse_line(line):
    name: str
    base_url_str: str
    tag_str: str
    doc_ext_str: str
    url_keyw_str: str
    collection_method: str
    scrape_method = "SimpleDocumentScrape"
    cron = "0 16 * * *"
    name, base_url_str, tag_str, doc_ext_str, url_keyw_str, collection_method = line  # type: ignore

    tags = tag_str.split(",") if tag_str else []
    base_urls = base_url_str.split("|") if base_url_str else []
    doc_exts = doc_ext_str.split(",") if doc_ext_str else ["pdf"]
    url_keyws = url_keyw_str.split(",") if url_keyw_str else []
    collection_method = collection_method if collection_method else "AUTOMATED"
    scrape_method_configuration = ScrapeMethodConfiguration(
        document_extensions=doc_exts,
        url_keywords=url_keyws,
        proxy_exclusions=[],
        wait_for=[],
        follow_links=False,
        follow_link_keywords=[],
        follow_link_url_keywords=[],
    )

    clean_urls = []
    for base_url in base_urls:
        try:
            parsed_url = BaseUrl(url=base_url)
            clean_urls.append(parsed_url)
        except ValidationError:
            logging.error(f"site {name} has invalid url: {base_url}")

    return Site(
        name=name,
        base_urls=clean_urls,
        scrape_method=scrape_method,
        collection_method=collection_method,
        scrape_method_configuration=scrape_method_configuration,
        tags=tags,
        disabled=False,
        cron=cron,
    )


def get_sites_from_json(file: UploadFile):
    content = file.file.read()
    content_obj = json.loads(content)
    for site in content_obj:
        site["disabled"] = False
        site["status"] = SiteStatus.NEW
        new_site = Site.parse_obj(site)
        yield new_site


def get_lines_from_xlsx(file: UploadFile):
    wb = load_workbook(io.BytesIO(file.file.read()))
    sheet = wb[wb.sheetnames[0]]

    for i, line in enumerate(sheet.values):
        # Skip header.
        if i == 0:
            continue
        # Skip blank site names. Happens with last line.
        if not line[0]:
            continue
        # Remove illegal characters.
        clean_line = []
        for line_value in line:
            if isinstance(line_value, str):
                clean_line.append(ILLEGAL_CHARACTERS_RE.sub("", line_value))
            else:
                clean_line.append(line_value)
        # Yield parsed site.
        yield parse_line(clean_line)


def get_lines_from_text_file(file: UploadFile):
    for line in file.file:
        line = line.decode("utf-8").strip()
        yield parse_line(line.split("\t"))


def get_sites_from_upload(file: UploadFile):
    if file.content_type == "application/json":
        return get_sites_from_json(file)
    try:
        return get_lines_from_xlsx(file)
    except zipfile.BadZipFile:
        return get_lines_from_text_file(file)


@router.post("/upload", response_model=list[Site])
async def upload_sites(
    file: UploadFile,
    current_user: User = Security(get_current_user),
    logger: Logger = Depends(get_logger),
):
    new_sites: list[Site] = []
    for new_site in get_sites_from_upload(file):
        if await Site.find_one(Site.base_urls == new_site.base_urls):
            continue
        if new_site.base_urls:
            new_sites.append(new_site)
            await create_and_log(logger, current_user, new_site)

    return new_sites


@router.post("/bulk-assign")
async def update_multiple_sites(
    updates: list[UpdateSiteAssigne],
    current_user: User = Security(get_current_user),
    logger: Logger = Depends(get_logger),
):
    # site_ids = map(lambda update: update.id, updates)
    site_ids = [update.id for update in updates]
    targets: list[Site] = await Site.find_many({"_id": {"$in": site_ids}}).to_list()

    result = []
    for target in targets:
        updated = await update_and_log_diff_forground(
            logger, current_user, target, UpdateSite(assignee=current_user.id)
        )
        result.append(updated)
    return result


@router.post("/{id}", response_model=Site)
async def update_site(
    updates: UpdateSite,
    target: Site = Depends(get_target),
    current_user: User = Security(get_current_user),
    logger: Logger = Depends(get_logger),
):
    updated = await update_and_log_diff(logger, current_user, target, updates)
    return updated


async def check_for_scrapetask(site_id: PydanticObjectId) -> list[SiteScrapeTask]:
    scrape_task = await SiteScrapeTask.find_many({"site_id": site_id}).to_list()
    return scrape_task


@router.delete(
    "/{id}",
    responses={
        405: {"description": "Item can't be deleted because of associated collection records."},
    },
)
async def delete_site(
    id: PydanticObjectId,
    target: Site = Depends(get_target),
    current_user: User = Security(get_current_user),
    logger: Logger = Depends(get_logger),
):
    # check for associated collection records, return error if present
    scrape_task = False  # await check_for_scrapetask(id) - reimplement check at later date.
    if scrape_task:
        raise HTTPException(
            status_code=status.HTTP_405_METHOD_NOT_ALLOWED,
            detail="Sites with run collections cannot be deleted.",
        )

    await update_and_log_diff(logger, current_user, target, UpdateSite(disabled=True))
    return {"success": True}
